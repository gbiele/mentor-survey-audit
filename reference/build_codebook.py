#!/usr/bin/env python3
"""Build canonical codebook from the FHI-EN questionnaire and export headers."""

from __future__ import annotations

import csv
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date
from html import escape
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT.parent / "data"
FHI_PATH = DATA_DIR / "mentor_fhi-EN.xlsx" if (DATA_DIR / "mentor_fhi-EN.xlsx").exists() else ROOT / "mentor_fhi-EN.xlsx"
EXP_PATH = DATA_DIR / "Content_Export_mentor_fhi_variabler_og_id.xlsx"
OUT_DIR = ROOT / "output"

META_KEYS = {"TITLE", "LOGOTEXT", "ESCAPEPAGE", "CONFIRMATIONPAGE"}
HEADER_RE = re.compile(r"^(.*?)\s*[\(\[]([^\)\]]+)[\)\]]\s*$")
NUMBERED_LABEL_RE = re.compile(r"^(\d+)\s*=")
PREFER_NOT_RE = re.compile(
    r"prefer not to (answer|say|respond|tell)|would rather not (answer|say)|"
    r"keine angabe",
    re.I,
)
DONT_KNOW_RE = re.compile(
    r"^(i )?((don't|dont|do not) know|dk|weiß nicht|weiss nicht|weiß es nicht)\.?$",
    re.I,
)
NOT_APPLICABLE_RE = re.compile(
    r"^(not applicable|n/?a|n\. ?a\.?|does not apply|not relevant|"
    r"nicht zutreffend|trifft nicht zu)\.?$",
    re.I,
)
NA_CODE = 997
DONT_KNOW_CODE = 998
PREFER_NOT_CODE = 999
SPECIAL_CODES = frozenset({NA_CODE, DONT_KNOW_CODE, PREFER_NOT_CODE})

FREQUENCY_RANKS: list[tuple[int, tuple[str, ...]]] = [
    (0, ("never true", "not at all", "never", "none")),
    (1, ("once in a while", "only on single days", "a few times", "rarely", "once")),
    (2, ("sometimes true", "half the time", "sometimes", "somewhat")),
    (3, ("often true", "quite often", "most of the time", "many times", "often")),
    (4, ("nearly all the time", "almost always", "very often", "always", "a lot")),
]
FREQ_HINTS = (
    "never", "rarely", "often", "always", "sometimes",
    "once in a while", "half the time", "many times", "a few times",
    "quite often", "very often", "almost always", "nearly all the time",
    "most of the time", "once",
)
UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", re.I
)

KNOWN_SCALES: list[list[str]] = [
    ["Yes", "No", "Don't know"],
    ["Yes", "No"],
    ["NEVER true", "SOMETIMES true", "OFTEN true"],
    ["Not at all", "A little", "Somewhat", "Quite a bit", "A lot"],
    ["Never", "Rarely", "Quite often", "Very often", "Always"],
    ["Not at all", "A little", "Quite", "Very", "To a great extent"],
    ["Less than 1 hour", "1-2 hours", "3-4 hours", "5-6 hours", "7 hours or more"],
    ["Nearly all the time", "Often", "Sometimes", "Rarely", "Never"],
    ["Always", "Most of the time", "Sometimes", "Rarely", "Never"],
    ["Many times", "A few times", "Once", "Never"],
    ["Strongly disagree", "Somewhat disagree", "Somewhat agree", "Strongly agree"],
    ["Bottom of the ladder", "Top of the ladder"],
]

SKIP_TEXT_PREFIXES = (
    "thank you",
    "do you wish to participate in a gift",
    "do you want to talk to someone",
    "background and purpose",
    "this survey has not yet been published",
)

NEW_QUESTION_RE = re.compile(
    r"^(please indicate|please select up to|when gaming|"
    r"how were you bullied most often|peer violence|"
    r"family environment|exposure to war|"
    r"this next question|these next questions|these questions are about)",
    re.I,
)

MULTIPLE_RE = re.compile(
    r"(which adults|select up to|what you usually use|"
    r"what do you identify as|please indicate what you usually)",
    re.I,
)

INTERVAL_VAR_HINTS = {"born1", "born3", "soc1"}
TEXT_VAR_HINTS = {"ID27"}

ORDINAL_SCALE_NORMS = {
    tuple(n.lower() for n in scale) for scale in KNOWN_SCALES if scale[0] != "Yes"
}
ORDINAL_SCALE_NORMS.add(("never true", "sometimes true", "often true"))
ORDINAL_SCALE_NORMS.add(
    ("less than high school", "high school diploma or equivalent",
     "come college, no degree", "bachelor's degree", "master's degree or more",
     "don't know")
)
ORDINAL_SCALE_NORMS.add(
    ("8th", "9th", "10th", "11th", "12th", "13th",
     "graduated from school", "quit school before graduating")
)
ORDINAL_SCALE_NORMS.add(("none", "1", "2", "3 or more"))


def norm(text: str | None) -> str:
    if text is None:
        return ""
    s = unicodedata.normalize("NFKC", str(text))
    s = s.replace("\xa0", " ").replace("&#xa0;", " ").replace("&nbsp;", " ")
    s = s.replace("–", "-").replace("—", "-").replace("…", "...")
    s = s.replace("’", "'").replace("‘", "'").replace("`", "'")
    s = re.sub(r"<[^>]+>", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\s+([.,;:])", r"\1", s)
    return s


def norm_key(text: str | None) -> str:
    return norm(text).casefold()


def slugify(text: str, fallback: str = "item") -> str:
    s = norm_key(text)
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return (s[:48] or fallback)


def looks_like_skip(text: str) -> bool:
    k = norm_key(text)
    return any(k.startswith(p) for p in SKIP_TEXT_PREFIXES)


def looks_like_new_question(text: str) -> bool:
    return bool(NEW_QUESTION_RE.match(norm(text)))


def strip_html_keep_text(text: str) -> str:
    return norm(text)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Option:
    label: str
    order: int
    value: int | None
    eusurvey_answer_id: str | None = None
    label_with_id: str | None = None
    aliases: list[str] = field(default_factory=list)


@dataclass
class Item:
    variable: str  # original export / EUSurvey variable name
    item_text: str
    source: str  # canonical_en
    ger_header: str = ""
    short_name: str = ""  # analysis name: ≤8 letter stem + item number
    is_core: bool = True


@dataclass
class Question:
    group_id: str
    section: str
    stem: str
    question_type: str  # single | matrix | text | interval
    scale: str
    scale_confidence: str
    multiple: bool
    options: list[Option]
    items: list[Item]
    options_complete: str  # true | false | inferred
    source: str
    is_core: bool = True
    notes: str = ""


@dataclass
class ExpVar:
    col: int
    vid: str
    header: str
    stem: str
    item: str
    is_mx: bool


@dataclass
class FhiRow:
    idx: int  # 1-based spreadsheet row
    uid: str
    raw: str
    text: str
    key: str


def parse_header(header: str) -> tuple[str, str]:
    header = header or ""
    m = HEADER_RE.match(header.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return header.strip(), "?"


def split_stem_item(text: str) -> tuple[str, str, bool]:
    raw = (text or "").strip()
    is_matrix = raw.startswith("Matrix:")
    if is_matrix:
        raw = raw[len("Matrix:"):].strip()
    if ":" in raw:
        stem, item = raw.rsplit(":", 1)
        stem, item = stem.strip(), item.strip()
        if not item:
            return stem or raw, "", is_matrix
        return stem, item, is_matrix
    if is_matrix:
        return "", raw, True
    return raw, "", False


def load_fhi_rows() -> list[FhiRow]:
    wb = load_workbook(FHI_PATH, data_only=True)
    ws = wb["Translation"]
    out: list[FhiRow] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        a, b, c = (list(row) + [None, None, None])[:3]
        uid = str(a or "")
        raw = "" if c is None else str(c)
        text = strip_html_keep_text(raw)
        out.append(FhiRow(idx=i, uid=uid, raw=raw, text=text, key=norm_key(text)))
    return out


def parse_export_headers() -> list[ExpVar]:
    wb = load_workbook(EXP_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(4, c).value for c in range(1, ws.max_column + 1) if ws.cell(4, c).value]
    vars_: list[ExpVar] = []
    for col, header in enumerate(headers, 1):
        text, vid = parse_header(str(header or ""))
        stem, item, is_mx = split_stem_item(text)
        vars_.append(
            ExpVar(
                col=col,
                vid=vid,
                header=text,
                stem=stem,
                item=item,
                is_mx=is_mx,
            )
        )
    return vars_


def match_known_scale(rows: list[FhiRow], i: int) -> list[str] | None:
    for scale in KNOWN_SCALES:
        end = i + len(scale)
        if end > len(rows):
            continue
        got = [rows[i + j].text for j in range(len(scale))]
        if got == scale:
            return scale
    return None


def parse_fhi(exp_stems: set[str], exp_questions: set[str]) -> tuple[list[dict], list[str]]:
    rows = load_fhi_rows()
    n = len(rows)

    section_at = {}
    current_section = "Background"
    stem_flags = [False] * n

    stem_keys = {norm_key(s) for s in exp_stems if s}
    q_keys = {norm_key(s) for s in exp_questions if s}

    for i, r in enumerate(rows):
        if i + 1 < n and rows[i + 1].text == "[Section]":
            current_section = r.text or current_section
            section_at[i] = current_section
            continue
        if r.uid in META_KEYS or r.text == "[Section]" or not r.text:
            continue
        if looks_like_skip(r.text):
            continue
        nxt_scale = match_known_scale(rows, i + 1) if i + 1 < n else None
        if nxt_scale:
            stem_flags[i] = True
        if r.key in stem_keys or r.key in q_keys:
            stem_flags[i] = True

    questions: list[dict] = []
    current_section = "Background"
    i = 0
    while i < n:
        r = rows[i]
        if i + 1 < n and rows[i + 1].text == "[Section]":
            current_section = r.text or current_section
            i += 2
            continue
        if r.uid in META_KEYS or r.text == "[Section]" or not r.text:
            i += 1
            continue
        if looks_like_skip(r.text):
            i += 1
            continue
        if not stem_flags[i] and not match_known_scale(rows, i + 1 if i + 1 < n else n):
            i += 1
            continue

        stem = r.text
        i += 1
        scale = match_known_scale(rows, i)
        options: list[str] = []
        items: list[str] = []
        if scale:
            options = list(scale)
            i += len(scale)
            while i < n:
                rr = rows[i]
                if rr.text == "[Section]":
                    break
                if i + 1 < n and rows[i + 1].text == "[Section]":
                    break
                if stem_flags[i]:
                    break
                if looks_like_skip(rr.text):
                    break
                if looks_like_new_question(rr.text) and items:
                    break
                if looks_like_new_question(rr.text) and not items and rr.key not in q_keys:
                    break
                if not rr.text:
                    i += 1
                    continue
                items.append(rr.text)
                i += 1
        else:
            while i < n:
                rr = rows[i]
                if rr.text == "[Section]":
                    break
                if i + 1 < n and rows[i + 1].text == "[Section]":
                    break
                if stem_flags[i]:
                    break
                if looks_like_skip(rr.text):
                    break
                if looks_like_new_question(rr.text) and options:
                    break
                if not rr.text:
                    i += 1
                    continue
                options.append(rr.text)
                i += 1

        questions.append(
            {
                "section": current_section,
                "stem": stem,
                "options": options,
                "items": items,
                "row": r.idx,
            }
        )

        while i < n and looks_like_new_question(rows[i].text) and not stem_flags[i]:
            ins = rows[i]
            if looks_like_skip(ins.text):
                break
            i += 1
            ins_opts: list[str] = []
            while i < n:
                rr = rows[i]
                if rr.text == "[Section]" or (i + 1 < n and rows[i + 1].text == "[Section]"):
                    break
                if stem_flags[i]:
                    break
                if looks_like_skip(rr.text) or looks_like_new_question(rr.text):
                    break
                if not rr.text:
                    i += 1
                    continue
                ins_opts.append(rr.text)
                i += 1
            questions.append(
                {
                    "section": current_section,
                    "stem": ins.text,
                    "options": ins_opts,
                    "items": [],
                    "row": ins.idx,
                }
            )

    return questions, []


def merge_empty_intro_questions(questions: list[dict]) -> list[dict]:
    out: list[dict] = []
    pending = ""
    for q in questions:
        empty = not q["options"] and not q["items"]
        if empty and not looks_like_skip(q["stem"]):
            pending = (pending + " " + q["stem"]).strip()
            continue
        if pending:
            q = dict(q)
            q["stem"] = f"{pending} {q['stem']}".strip()
            pending = ""
        out.append(q)
    if pending:
        out.append(
            {
                "section": questions[-1]["section"] if questions else "",
                "stem": pending,
                "options": [],
                "items": [],
                "row": 0,
            }
        )
    return out


# ---------------------------------------------------------------------------
# Scale typing and codes
# ---------------------------------------------------------------------------

def is_prefer_not(label: str) -> bool:
    return bool(PREFER_NOT_RE.search(norm(label)))


def special_missing_code(label: str) -> int | None:
    text = norm(label)
    if not text:
        return None
    if is_prefer_not(text):
        return PREFER_NOT_CODE
    if DONT_KNOW_RE.match(text):
        return DONT_KNOW_CODE
    if NOT_APPLICABLE_RE.match(text):
        return NA_CODE
    return None


def is_special_missing(label: str) -> bool:
    return special_missing_code(label) is not None


def frequency_rank(label: str) -> int | None:
    k = norm_key(label)
    for rank, phrases in FREQUENCY_RANKS:
        for p in phrases:
            if k == p or k.startswith(p + " ") or k.endswith(" " + p):
                return rank
    return None


def is_frequency_scale(labels: list[str]) -> bool:
    keys = [norm_key(x) for x in labels if not is_special_missing(x)]
    if len(keys) < 2:
        return False
    blob = " ".join(keys)
    return any(h in blob for h in FREQ_HINTS)


def reverse_frequency_values(labels: list[str]) -> bool:
    if not is_frequency_scale(labels):
        return False
    ranks = [frequency_rank(x) for x in labels if not is_special_missing(x)]
    ranks = [r for r in ranks if r is not None]
    if len(ranks) < 2:
        return False
    return ranks[0] > ranks[-1]


def option_value(label: str, order: int, scale: str) -> int | None:
    special = special_missing_code(label)
    if special is not None:
        return special
    m = NUMBERED_LABEL_RE.match(label.strip())
    if m:
        return int(m.group(1))
    if re.fullmatch(r"-?\d+", label.strip()):
        return int(label.strip())
    if scale == "text":
        return None
    if scale == "interval":
        if label.lower().startswith("bottom"):
            return 1
        if label.lower().startswith("top"):
            return 10
        return None
    if scale == "ordinal":
        return order
    if scale == "nominal":
        return order + 1
    return order


def classify_scale(
    stem: str,
    items: list[str],
    options: list[str],
    variables: list[str],
) -> tuple[str, str, str, str]:
    vars_l = {v for v in variables}
    stem_k = norm_key(stem)
    opt_tuple = tuple(norm_key(o) for o in options)
    opt_set = set(opt_tuple)

    if any(v in TEXT_VAR_HINTS for v in vars_l) or stem_k in {"describe", "describe:"}:
        return "text", "text", "high", "true"
    if not options and not any(v in INTERVAL_VAR_HINTS for v in vars_l):
        if any("describe" in (it or "").lower() for it in items + [stem]):
            return "text", "text", "high", "true"
        return "nominal", "single", "medium", "false"

    if any(v in INTERVAL_VAR_HINTS for v in vars_l) or opt_tuple == (
        "bottom of the ladder",
        "top of the ladder",
    ):
        return "interval", "interval", "high", "true"

    ints = []
    if options and all(re.fullmatch(r"-?\d+", o.strip()) for o in options):
        ints = [int(o.strip()) for o in options]
        if ints == list(range(ints[0], ints[0] + len(ints))):
            return "interval", "interval", "high", "true"

    if opt_tuple in ORDINAL_SCALE_NORMS or opt_tuple[::-1] in ORDINAL_SCALE_NORMS:
        qtype = "matrix" if len(items) > 1 else "single"
        return "ordinal", qtype, "high", "true"

    months = {
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december",
    }
    if opt_set <= months and len(opt_set) >= 10:
        return "ordinal", "single", "high", "true"

    if "education" in stem_k or "grade are you in" in stem_k or "siblings" in stem_k:
        return "ordinal", "single", "high", "true"

    core_opts = {k for k in opt_set if special_missing_code(k) is None}
    yes_no = core_opts <= {"yes", "no"} and "yes" in core_opts and "no" in core_opts
    if yes_no:
        qtype = "matrix" if len(items) > 1 else "single"
        return "nominal", qtype, "high", "true"

    joined = " | ".join(opt_tuple)
    if any(w in joined for w in (
        "agree", "disagree", "never", "often", "rarely", "always",
        "somewhat", "strongly", "difficult", "einfach", "schwierig",
        "not at all", "quite a bit", "once in a while", "half the time",
    )):
        qtype = "matrix" if len(items) > 1 else "single"
        return "ordinal", qtype, "high", "true"

    if len(items) > 1:
        return "nominal", "matrix", "medium", "true"
    return "nominal", "single", "high", "true"


def is_multiple(stem: str, item_text: str = "") -> bool:
    blob = f"{stem} {item_text}"
    return bool(MULTIPLE_RE.search(blob))


def make_options(
    labels: list[str],
    scale: str,
    id_map: dict[str, str] | None = None,
    aliases: dict[str, list[str]] | None = None,
) -> list[Option]:
    id_map = id_map or {}
    aliases = aliases or {}
    out: list[Option] = []
    month_order = [
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december",
    ]
    is_months = {norm_key(x) for x in labels} <= set(month_order) and len(labels) >= 10
    core_keys = {norm_key(x) for x in labels if not is_special_missing(x)}
    is_yes_no = core_keys == {"yes", "no"}
    reverse_freq = scale == "ordinal" and reverse_frequency_values(labels)
    freq_idx = [
        i for i, lab in enumerate(labels)
        if frequency_rank(lab) is not None and not is_special_missing(lab)
    ]
    n_freq = len(freq_idx)
    freq_pos = {i: j for j, i in enumerate(freq_idx)}
    regular_i = 0

    for order, lab in enumerate(labels):
        special = special_missing_code(lab)
        if special is not None:
            value = special
        elif is_yes_no:
            value = 0 if norm_key(lab) == "no" else 1
        elif is_months:
            value = month_order.index(norm_key(lab)) + 1 if norm_key(lab) in month_order else order + 1
        elif scale == "interval" and re.fullmatch(r"-?\d+", lab.strip()):
            value = int(lab.strip())
        elif reverse_freq and order in freq_pos:
            value = n_freq - 1 - freq_pos[order]
        else:
            value = option_value(lab, regular_i, scale)
            regular_i += 1
        aid = ""
        if lab in id_map:
            aid = id_map[lab]
        else:
            for k, v in id_map.items():
                if norm_key(k) == norm_key(lab):
                    aid = v
                    break
        with_id = f"{lab} ({aid})" if aid else ""
        als = list(aliases.get(lab, []))
        out.append(
            Option(
                label=lab,
                order=order,
                value=value,
                eusurvey_answer_id=aid or None,
                label_with_id=with_id or None,
                aliases=als,
            )
        )
    return out


def unique_group_id(used: set[str], stem: str, fallback: str) -> str:
    base = slugify(stem, fallback)
    gid = base
    n = 2
    while gid in used:
        gid = f"{base}_{n}"
        n += 1
    used.add(gid)
    return gid


def propose_stem(q: Question) -> str:
    first = q.items[0].variable if q.items else ""
    stem_k = norm_key(q.stem)
    item0 = norm_key(q.items[0].item_text if q.items else "")
    sec = q.section

    if "wish to participate" in stem_k:
        return "consent"
    if "what gender do you identify" in stem_k:
        return "gender"
    if "what do you identify as" in stem_k:
        return "ident"
    if "what grade are you in" in stem_k:
        return "grade"
    if "what year were you born" in stem_k:
        return "byear"
    if "what month were you born" in stem_k:
        return "bmonth"
    if "where were you born" in stem_k and "mother" not in stem_k and "father" not in stem_k:
        return "bplace"
    if "how old were you when you came" in stem_k:
        return "bage"
    if "mother born" in stem_k:
        return "bmborn"
    if "father born" in stem_k:
        return "bfborn"
    if "thinking about the people you live with" in stem_k:
        return "live"
    if "which adults do you live with" in stem_k:
        return "livead"
    if "what best describes your situation" in stem_k:
        return "livesp"
    if "mother have a new partner" in stem_k:
        return "momprt"
    if "father have a new partner" in stem_k:
        return "dadprt"
    if "siblings" in stem_k:
        return "sibs"
    if "mother" in stem_k and "education" in stem_k:
        return "edumom"
    if "father" in stem_k and "education" in stem_k:
        return "edudad"
    if "imagine a ladder" in stem_k:
        return "ladder"
    if "usually afford" in stem_k:
        return "afford"

    if sec == "Mental health":
        m = re.fullmatch(r"([a-z][a-z_]{0,7})\d+$", first.lower())
        if m and m.group(1) != "id":
            return f"bcfpi_{m.group(1)}"

    if sec.startswith("Adverse"):
        # Match variable prefix if available (e.g. ace, svr)
        m = re.fullmatch(r"([a-z][a-z_]{0,7})\d+$", first.lower())
        if m and m.group(1) != "id":
            return m.group(1)
        return "ace"

    if "to what extent do the following" in stem_k or first.startswith("cyrm"):
        return "cyrm"
    if sec == "Quality of life" or first.startswith("ks"):
        return "ks"
    if "normal weekday" in stem_k or first.startswith("some_week"):
        return "smwd"
    if "normal weekend" in stem_k or first.startswith("some_weekend"):
        return "smwe"
    if "when gaming on your console" in stem_k or first.startswith("gaming"):
        return "gaming"
    if "do you ever use ai" in stem_k:
        return "ai"
    if "usually use ai" in stem_k or first == "ai2":
        return "aiuse"
    if "positive feelings after" in stem_k or first == "some_feel1":
        return "smpos"
    if "negative feelings after" in stem_k or first == "some_feel2":
        return "smneg"
    if "negative effect" in stem_k or first.startswith("some_use"):
        return "smharm"
    if "apps or services" in stem_k or first.startswith("some_apps"):
        return "apps"
    if first == "ID27":
        return "draw"

    return _abbrev_stem(q.stem)


def _abbrev_stem(text: str) -> str:
    words = re.findall(r"[a-z]+", norm_key(text))
    stop = {
        "the", "a", "an", "of", "and", "or", "to", "in", "your", "you", "do",
        "did", "how", "what", "when", "for", "with", "that", "this", "are",
        "is", "was", "were", "best", "please", "select",
    }
    keep = [w for w in words if w not in stop] or words
    acro = "".join(w[0] for w in keep)[:8]
    if len(acro) >= 3:
        return acro
    return (keep[0] if keep else "var")[:8]


STEM_MAX = 10
SHORT_NAME_RE = re.compile(r"^(?:bcfpi_[a-z]+|[a-z_]{1,10})\d+$")


def _unique_stem(base: str, used: set[str]) -> str:
    raw = re.sub(r"[^a-z_]", "", (base or "var").lower()) or "var"
    if raw.startswith("bcfpi_"):
        s = raw
        limit = None
    else:
        s = raw[:STEM_MAX]
        limit = STEM_MAX
    if s not in used:
        return s
    for ch in "abcdefghijklmnopqrstuvwxyz":
        if limit is None:
            cand = s + ch
        else:
            cand = (s[: limit - 1] + ch) if len(s) >= limit else s + ch
        if cand not in used:
            return cand
    raise RuntimeError(f"Could not uniquify stem {base!r}")


def assign_short_names(codebook: list[Question]) -> None:
    used_stems: set[str] = set()
    used_names: set[str] = set()
    for q in codebook:
        stem = _unique_stem(propose_stem(q), used_stems)
        used_stems.add(stem)
        n = len(q.items)
        width = 2 if n >= 10 else 1
        for i, it in enumerate(q.items, 1):
            name = f"{stem}{i:0{width}d}"
            if name in used_names:
                raise RuntimeError(f"Duplicate short name {name} for {it.variable}")
            if not SHORT_NAME_RE.match(name):
                raise RuntimeError(f"Invalid short name {name!r}")
            it.short_name = name
            used_names.add(name)


def short_name_range(q: Question) -> str:
    names = [it.short_name or it.variable for it in q.items]
    if not names:
        return ""
    if len(names) == 1:
        return names[0]
    return f"{names[0]}–{names[-1]}"


# ---------------------------------------------------------------------------
# Codebook Builder (Canonical FHI Baseline)
# ---------------------------------------------------------------------------

def build_codebook() -> list[Question]:
    exp_vars = parse_export_headers()
    exp_stems = {ev.stem for ev in exp_vars if ev.stem}
    exp_questions = {ev.header for ev in exp_vars}
    for ev in exp_vars:
        if ev.stem and not ev.item:
            exp_questions.add(ev.stem)

    fhi_qs, _ = parse_fhi(exp_stems, exp_questions)
    fhi_qs = merge_empty_intro_questions(fhi_qs)

    used_exp_cols: set[int] = set()
    used_gids: set[str] = set()
    codebook: list[Question] = []

    for qi, fq in enumerate(fhi_qs):
        stem = fq["stem"]
        stem_k = norm_key(stem)
        items: list[Item] = []

        if fq["items"]:
            for it in fq["items"]:
                it_k = norm_key(it)
                matched_ev = None
                for ev in exp_vars:
                    if ev.col in used_exp_cols:
                        continue
                    ev_it_k = norm_key(ev.item)
                    ev_stem_k = norm_key(ev.stem)
                    if ev_it_k and ev_it_k == it_k:
                        if not ev_stem_k or ev_stem_k == stem_k or ev_stem_k in stem_k or stem_k in ev_stem_k:
                            matched_ev = ev
                            break
                        elif len(it_k) >= 15:
                            matched_ev = ev
                            break

                if matched_ev:
                    used_exp_cols.add(matched_ev.col)
                    items.append(
                        Item(
                            variable=matched_ev.vid,
                            item_text=it,
                            source="canonical_en",
                            ger_header=matched_ev.header,
                        )
                    )
                else:
                    vid = f"fhi_{qi+1}_{len(items)+1}"
                    items.append(Item(variable=vid, item_text=it, source="canonical_en"))
        else:
            matched_ev = None
            for ev in exp_vars:
                if ev.col in used_exp_cols:
                    continue
                ev_hdr_k = norm_key(ev.header)
                ev_stem_k = norm_key(ev.stem)
                if ev_hdr_k == stem_k or ev_stem_k == stem_k:
                    matched_ev = ev
                    break
                elif stem_k in ev_hdr_k or ev_hdr_k in stem_k:
                    matched_ev = ev
                    break

            if matched_ev:
                used_exp_cols.add(matched_ev.col)
                items.append(
                    Item(
                        variable=matched_ev.vid,
                        item_text="",
                        source="canonical_en",
                        ger_header=matched_ev.header,
                    )
                )

        if not items:
            continue

        variables = [it.variable for it in items]
        scale, qtype, conf, _ = classify_scale(
            fq["stem"], [it.item_text for it in items], fq["options"], variables
        )
        if len(items) > 1:
            qtype = "matrix"
        elif scale == "interval":
            qtype = "interval"
        elif scale == "text":
            qtype = "text"
        else:
            qtype = "single"

        if qtype == "interval" and fq["options"] == ["Bottom of the ladder", "Top of the ladder"]:
            labels = [str(i) for i in range(1, 11)]
            notes = "MacArthur ladder 1 (bottom) to 10 (top). FHI only stores endpoint labels."
            opts = make_options(labels, "interval")
            opts[0].label = "1"
            opts[0].aliases = ["Bottom of the ladder"]
            opts[-1].label = "10"
            opts[-1].aliases = ["Top of the ladder"]
            options_complete = "true"
        elif "what grade are you in" in norm_key(fq["stem"]):
            notes = (
                "Options 'graduated from school' (code 6) and 'quit school before graduating' "
                "(code 7) are terminal non-enrolled statuses; exclude or handle separately when "
                "treating grade level as continuous/ordinal."
            )
            opts = make_options(fq["options"], scale)
            options_complete = "true" if fq["options"] else ("true" if scale == "text" else "false")
        else:
            notes = ""
            opts = make_options(fq["options"], scale)
            options_complete = "true" if fq["options"] else ("true" if scale == "text" else "false")

        multiple = is_multiple(fq["stem"]) or any(
            is_multiple(fq["stem"], it.item_text) for it in items
        )
        gid = unique_group_id(used_gids, fq["stem"], f"q{qi}")
        codebook.append(
            Question(
                group_id=gid,
                section=fq["section"],
                stem=fq["stem"],
                question_type=qtype,
                scale=scale,
                scale_confidence=conf,
                multiple=multiple,
                options=opts,
                items=items,
                options_complete=options_complete,
                source="canonical_en",
                notes=notes,
            )
        )

    # Remaining FHI export variables (e.g. ID27 at the end)
    remaining_exp = [ev for ev in exp_vars if ev.col not in used_exp_cols]
    for ev in remaining_exp:
        gid = unique_group_id(used_gids, ev.header, f"exp_{ev.vid}")
        codebook.append(
            Question(
                group_id=gid,
                section="Social media and gaming",
                stem=ev.header,
                question_type="text",
                scale="text",
                scale_confidence="high",
                multiple=False,
                options=[],
                items=[Item(variable=ev.vid, item_text="", source="canonical_en", ger_header=ev.header)],
                options_complete="true",
                source="canonical_en",
                notes="Free text response / SMS code",
            )
        )
        used_exp_cols.add(ev.col)

    assign_short_names(codebook)
    return codebook


# ---------------------------------------------------------------------------
# CSV & HTML writers
# ---------------------------------------------------------------------------

def write_csv(codebook: list[Question]) -> None:
    OUT_DIR.mkdir(exist_ok=True, parents=True)
    var_path = OUT_DIR / "codebook_variables.csv"
    opt_path = OUT_DIR / "codebook_options.csv"

    with var_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "variable",
                "orig_variable",
                "group_id",
                "section",
                "question_stem",
                "item_text",
                "scale",
                "question_type",
                "n_options",
                "multiple",
                "source",
                "is_core",
                "options_complete",
                "scale_confidence",
                "notes",
            ],
        )
        w.writeheader()
        for q in codebook:
            for it in q.items:
                w.writerow(
                    {
                        "variable": it.short_name or it.variable,
                        "orig_variable": it.variable,
                        "group_id": q.group_id,
                        "section": q.section,
                        "question_stem": q.stem,
                        "item_text": it.item_text,
                        "scale": q.scale,
                        "question_type": q.question_type,
                        "n_options": len(q.options),
                        "multiple": str(q.multiple).lower(),
                        "source": it.source,
                        "is_core": str(it.is_core).lower(),
                        "options_complete": q.options_complete,
                        "scale_confidence": q.scale_confidence,
                        "notes": q.notes,
                    }
                )

    with opt_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "variable",
                "orig_variable",
                "group_id",
                "option_order",
                "option_label",
                "option_label_with_id",
                "option_alias",
                "eusurvey_answer_id",
                "value",
                "scale",
                "multiple",
                "source",
                "options_complete",
            ],
        )
        w.writeheader()
        for q in codebook:
            for it in q.items:
                if not q.options:
                    continue
                for opt in q.options:
                    labels_to_write = [(opt.label, "")]
                    for a in opt.aliases:
                        if a and norm_key(a) != norm_key(opt.label):
                            labels_to_write.append((a, a))
                    for lab, alias in labels_to_write:
                        with_id = opt.label_with_id
                        if alias and opt.eusurvey_answer_id:
                            with_id = f"{lab} ({opt.eusurvey_answer_id})"
                        elif alias:
                            with_id = ""
                        w.writerow(
                            {
                                "variable": it.short_name or it.variable,
                                "orig_variable": it.variable,
                                "group_id": q.group_id,
                                "option_order": opt.order,
                                "option_label": lab,
                                "option_label_with_id": with_id or "",
                                "option_alias": alias,
                                "eusurvey_answer_id": opt.eusurvey_answer_id or "",
                                "value": "" if opt.value is None else opt.value,
                                "scale": q.scale,
                                "multiple": str(q.multiple).lower(),
                                "source": it.source,
                                "options_complete": q.options_complete,
                            }
                        )


def main() -> None:
    codebook = build_codebook()
    write_csv(codebook)
    print(f"Successfully wrote {OUT_DIR / 'codebook_variables.csv'}")
    print(f"Successfully wrote {OUT_DIR / 'codebook_options.csv'}")
    print(f"Total Questions: {len(codebook)}, Total Variables: {sum(len(q.items) for q in codebook)}")


if __name__ == "__main__":
    main()
