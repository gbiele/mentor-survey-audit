#!/usr/bin/env python3
"""Generate a synthetic survey test-data export with 50 respondents.

Preserves the exact 174-column header structure, the structural
missing/incomplete option characteristics, and introduces 3% to 6%
random item non-response (missing data) across all survey questions.
"""

import json
import random
from datetime import datetime
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
SRC_DIR = ROOT / "src"
TEMPLATE_FILE = DATA_DIR / "Content_Export_MENTORMasterGER1_Test-GER-1.xlsx"
DICT_FILE = DATA_DIR / "master_dictionary.json"

random.seed(42)  # Deterministic generation


def generate_synthetic_data(n_respondents=100):
    with open(DICT_FILE, encoding="utf-8") as f:
        master_dict = json.load(f)

    # Load template headers
    wb_template = openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)
    ws_template = wb_template[wb_template.sheetnames[0]]
    template_rows = list(ws_template.iter_rows(values_only=True))

    headers = list(template_rows[3])
    n_cols = len(headers)

    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Content"

    # Set metadata rows
    ws.append(["Alias", "MENTORMasterGER1_Synthetic100"] + [None] * (n_cols - 2))
    ws.append(["Export Date", datetime(2026, 8, 18, 17, 30, 0)] + [None] * (n_cols - 2))
    ws.append([None] * n_cols)
    ws.append(headers)

    # Map headers to dictionary variables
    from survey_validator_py import build_column_generators
    col_generators = build_column_generators(headers, master_dict, template_rows[4:])

    # Generate complete matrix of responses
    column_data = []
    for col_idx in range(n_cols):
        gen = col_generators[col_idx]
        col_vals = [gen(resp_id) for resp_id in range(1, n_respondents + 1)]

        # Introduce 3% to 6% missingness per question
        # For n=100: 3 to 6 missing values per column (3.0% to 6.0%)
        non_none_indices = [i for i, v in enumerate(col_vals) if v is not None and str(v).strip() != ""]
        if len(non_none_indices) > 20:
            n_missing = random.randint(3, 6)  # 3/100 = 3%, 6/100 = 6%
            missing_indices = random.sample(non_none_indices, min(n_missing, len(non_none_indices)))
            for idx in missing_indices:
                col_vals[idx] = None

        column_data.append(col_vals)

    # Write rows to worksheet
    for resp_idx in range(n_respondents):
        row_values = [column_data[col_idx][resp_idx] for col_idx in range(n_cols)]
        ws.append(row_values)

    out_file = DATA_DIR / "Content_Export_MENTORMasterGER1_Test-GER-1.xlsx"
    wb.save(out_file)
    print(f"Saved synthetic survey export with {n_respondents} respondents (3-6% missingness per item) to {out_file}")

    # Also update src/sample_data.js and docs/src/sample_data.js
    rows = list(ws.iter_rows(values_only=True))
    clean_rows = []
    for r in rows:
        clean_r = []
        for cell in r:
            if cell is None:
                clean_r.append(None)
            elif hasattr(cell, "isoformat"):
                clean_r.append(cell.isoformat())
            else:
                clean_r.append(cell)
        clean_rows.append(clean_r)

    update_sample_data_js(clean_rows)


def update_sample_data_js(clean_rows):
    js_file = SRC_DIR / "sample_data.js"
    docs_js_file = ROOT / "docs" / "src" / "sample_data.js"

    with open(js_file, "r", encoding="utf-8") as f:
        code = f.read()

    prefix = code.split("window.__BUNDLED_SAMPLE_ROWS__ =")[0]
    new_code = prefix + f"window.__BUNDLED_SAMPLE_ROWS__ = {json.dumps(clean_rows, ensure_ascii=False)};\n"

    with open(js_file, "w", encoding="utf-8") as f:
        f.write(new_code)

    if docs_js_file.parent.exists():
        with open(docs_js_file, "w", encoding="utf-8") as f:
            f.write(new_code)

    print(f"Updated sample_data.js with {len(clean_rows)} rows (50 respondents with 3-6% missingness)")


if __name__ == "__main__":
    generate_synthetic_data(100)
